from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def write_xlsx(
    path: Path,
    rows: list[dict[str, Any]],
    fields: list[str],
    sheet_name: str = "Rows",
    highlight_terms: list[str] | None = None,
    highlight_color: str = "FFF2CC",
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = sheet_name[:31]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    link_font = Font(color="0563C1", underline="single")
    terms = [term.strip().lower() for term in highlight_terms or [] if term.strip()]
    highlight_fill = PatternFill(
        start_color=normalize_fill_color(highlight_color),
        end_color=normalize_fill_color(highlight_color),
        fill_type="solid",
    )

    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, record in enumerate(rows, 2):
        for col_idx, field in enumerate(fields, 1):
            value = record.get(field, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if isinstance(value, str) and value.startswith(("http://", "https://")):
                cell.hyperlink = value
                cell.font = link_font
            if terms and any(term in str(value).lower() for term in terms):
                cell.fill = highlight_fill

    for col_idx, field in enumerate(fields, 1):
        max_len = len(field)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=len(rows) + 1):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 60)

    wb.save(path)


def normalize_fill_color(value: str) -> str:
    color = (value or "").strip().lstrip("#").upper()
    if len(color) == 8 and all(char in "0123456789ABCDEF" for char in color):
        return color
    if len(color) == 6 and all(char in "0123456789ABCDEF" for char in color):
        return color
    return "FFF2CC"
